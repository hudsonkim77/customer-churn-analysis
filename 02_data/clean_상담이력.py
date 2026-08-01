"""
정제_규칙_상담이력.md에 정리한 5가지 규칙을 그대로 구현한 정제 스크립트.

사용법:
    python clean_상담이력.py <입력 엑셀 경로>

규칙 (정제_규칙_상담이력.md 순서와 동일):
    1. 소속(A열) 병합 셀 forward-fill
    2. 소계 행을 개인 상담 테이블 / 팀 요약 테이블로 분리
    3. 상담일: 표준(YYYY-MM-DD)/점표기(YY.M.D)/엑셀 일련번호를 모두 datetime으로 통일
    4. 상담시간: "N분" 문자열과 숫자를 모두 숫자형으로 통일
    5. 재문의여부: 배경색(빨강=재문의)을 기준으로 True/False 변환

원본 엑셀 파일은 읽기만 하며 절대 수정하지 않는다.
"""

import argparse
import sys
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

EXCEL_EPOCH = date(1899, 12, 30)
RECONTACT_FILL = "FFFF0000"


def parse_consult_date(value):
    """규칙 3: 상담일 형식 통일 - 표준/점표기/엑셀 일련번호를 모두 date로 변환."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return EXCEL_EPOCH + timedelta(days=int(value))
    s = str(value).strip()
    if "." in s:
        year_str, month_str, day_str = s.split(".")
        return date(2000 + int(year_str), int(month_str), int(day_str))
    return pd.to_datetime(s).date()


def parse_duration(value):
    """규칙 4: 상담시간 형식 통일 - "N분" 문자열과 숫자를 모두 숫자로 변환."""
    if value is None:
        return None
    if isinstance(value, str):
        return int(value.replace("분", "").strip())
    return value


def parse_recontact(cell):
    """규칙 5: 재문의여부를 배경색(빨강) 기준으로 True/False 변환."""
    fill = cell.fill
    fg = fill.fgColor.rgb if fill and fill.patternType else None
    if fg == RECONTACT_FILL:
        return True
    if cell.value == "N":
        return False
    raise ValueError(f"예상치 못한 재문의여부 값: value={cell.value!r}, fill={fg!r}")


def clean_sheet(ws, year_month):
    """규칙 1·2: 병합 셀을 forward-fill하며, 소계 행을 개인/팀 요약 테이블로 분리."""
    individual_rows = []
    summary_rows = []
    current_team = None

    for row_idx in range(3, ws.max_row + 1):
        team_cell = ws.cell(row=row_idx, column=1).value
        agent = ws.cell(row=row_idx, column=2).value

        if team_cell is not None:
            if "소계" in str(team_cell):
                summary_rows.append({
                    "년월": year_month,
                    "소속": str(team_cell).replace(" 소계", ""),
                    "상담건수": int(str(ws.cell(row=row_idx, column=3).value).replace("건", "")),
                    "상담시간_합계": ws.cell(row=row_idx, column=7).value,
                })
                continue
            current_team = team_cell

        individual_rows.append({
            "년월": year_month,
            "소속": current_team,
            "담당자": agent,
            "고객번호": ws.cell(row=row_idx, column=3).value,
            "상담일": parse_consult_date(ws.cell(row=row_idx, column=4).value),
            "채널": ws.cell(row=row_idx, column=5).value,
            "문의유형": ws.cell(row=row_idx, column=6).value,
            "상담시간": parse_duration(ws.cell(row=row_idx, column=7).value),
            "처리결과": ws.cell(row=row_idx, column=8).value,
            "재문의여부": parse_recontact(ws.cell(row=row_idx, column=9)),
        })

    return pd.DataFrame(individual_rows), pd.DataFrame(summary_rows)


def clean(input_path: Path) -> tuple[pd.DataFrame, pd.DataFrame]:
    wb = openpyxl.load_workbook(input_path, data_only=False)
    individual_parts, summary_parts = [], []
    for sheet_name in wb.sheetnames:
        ind, summ = clean_sheet(wb[sheet_name], sheet_name)
        individual_parts.append(ind)
        summary_parts.append(summ)
    individual = pd.concat(individual_parts, ignore_index=True)
    summary = pd.concat(summary_parts, ignore_index=True)
    return individual, summary


def verify_duration_totals(individual: pd.DataFrame, summary: pd.DataFrame) -> pd.DataFrame:
    """검산: 개인 테이블의 상담시간 합계(년월·소속별)가 팀 요약의 상담시간_합계와 일치하는지 확인."""
    recomputed = individual.groupby(["년월", "소속"])["상담시간"].sum(min_count=1).reset_index()
    merged = summary.merge(recomputed, on=["년월", "소속"], suffixes=("_요약", "_재계산"))
    return merged[merged["상담시간_합계"] != merged["상담시간"]]


def main() -> None:
    parser = argparse.ArgumentParser(description="상담이력 messy 엑셀 정제 스크립트")
    parser.add_argument("input_path", type=Path, help="정제할 입력 엑셀(.xlsx) 파일 경로")
    args = parser.parse_args()

    input_path: Path = args.input_path
    if not input_path.exists():
        sys.exit(f"입력 파일을 찾을 수 없습니다: {input_path}")

    individual, summary = clean(input_path)
    mismatch = verify_duration_totals(individual, summary)

    stem = input_path.stem
    out_dir = input_path.parent
    individual_path = out_dir / f"{stem}_clean.csv"
    summary_path = out_dir / f"{stem}_팀요약.csv"

    for out_path in (individual_path, summary_path):
        if out_path.resolve() == input_path.resolve():
            sys.exit(f"출력 경로가 입력 파일과 같습니다. 중단합니다: {out_path}")

    individual.to_csv(individual_path, index=False, encoding="utf-8-sig")
    summary.to_csv(summary_path, index=False, encoding="utf-8-sig")

    print("=== 정제 결과 ===")
    print(f"개인 상담 테이블 행 수: {len(individual)}")
    print(f"팀 요약 테이블 행 수: {len(summary)}")
    print(f"재문의여부 분포: {individual['재문의여부'].value_counts().to_dict()}")
    print(f"상담시간 결측: {individual['상담시간'].isna().sum()}건")
    if len(mismatch):
        print(f"[경고] 상담시간 합계 불일치 {len(mismatch)}건 발견")
        print(mismatch.to_string(index=False))
    else:
        print("[확인] 상담시간 합계 검증: 팀 요약과 개인 테이블 재계산 값 전부 일치")
    print()
    print(f"저장: {individual_path}")
    print(f"저장: {summary_path}")


if __name__ == "__main__":
    main()
